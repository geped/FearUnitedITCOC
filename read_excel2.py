import sys, subprocess
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\pedro\Desktop\Claude\FearUnitedCoC\Bonus CWL - FearUnitedIT.xlsx')
ws = wb.active

print(f"Fogli: {wb.sheetnames}")
print(f"Righe: {ws.max_row}, Colonne: {ws.max_column}")

# Stampa intestazioni riga 1
row1 = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
print("\nRiga 1 (intestazioni):", row1)

# Stampa prime 3 righe di dati
print("\nPrime righe:")
for r in range(1, min(ws.max_row+1, 50)):
    row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
    if any(v is not None for v in row_vals):
        print(f"  R{r}: {row_vals}")
