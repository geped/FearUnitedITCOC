import subprocess, sys
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'])
import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\pedro\Desktop\Claude\FearUnitedCoC\Bonus CWL - FearUnitedIT.xlsx')
print('Fogli:', wb.sheetnames)
for sname in wb.sheetnames:
    ws = wb[sname]
    print(f'\n=== FOGLIO: {sname} ({ws.max_row} righe x {ws.max_column} col) ===')
    for row in ws.iter_rows(max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            print(list(row))
