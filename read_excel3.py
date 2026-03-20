import sys, subprocess, json, datetime
subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl', '-q'],
                       stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\pedro\Desktop\Claude\FearUnitedCoC\Bonus CWL - FearUnitedIT.xlsx')
ws = wb.active

out = {}
out['sheets'] = wb.sheetnames
out['rows'] = ws.max_row
out['cols'] = ws.max_column

# Leggi riga 1 = intestazioni
headers = []
for c in range(1, ws.max_column+1):
    v = ws.cell(1, c).value
    if isinstance(v, datetime.datetime):
        headers.append(v.strftime('%Y-%m'))
    else:
        headers.append(str(v) if v is not None else None)
out['headers'] = headers

# Leggi tutte le righe di dati (dalla riga 2 in poi)
data = []
for r in range(2, ws.max_row+1):
    row = []
    for c in range(1, ws.max_column+1):
        v = ws.cell(r, c).value
        if isinstance(v, datetime.datetime):
            row.append(v.strftime('%Y-%m'))
        else:
            row.append(v)
    if any(x is not None for x in row):
        data.append(row)
out['data'] = data

with open(r'C:\Users\pedro\Desktop\Claude\FearUnitedCoC\excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(out, f, ensure_ascii=False, indent=2, default=str)

print("OK - salvato in excel_data.json")
print(f"Headers: {headers}")
print(f"Righe dati: {len(data)}")
